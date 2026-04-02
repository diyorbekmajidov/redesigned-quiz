# ============================================
# ADMIN PANEL - PSIXOLOGIK VA STANDART TESTLAR
# ============================================

from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from .models import (
    Quiz, Question, QuestionText, Option,
    QuizAttempt, UserResponse, Result,
    PsychologicalScale, PsychologicalCategory,
    PsychologicalResult, PsychologicalScaleResult
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ==================== HELPERS ====================

def _make_excel_response(wb, filename):
    """Excel faylni HTTP response sifatida qaytarish"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _style_header_row(ws, row, num_cols, fill_color="4F46E5"):
    """Header qatorini stilizatsiya qilish"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _auto_column_width(ws):
    """Ustun kengliklarini avtomatik sozlash"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


COLOR_MAP = {
    'green': 'C6EFCE',
    'yellow': 'FFEB9C',
    'orange': 'FFCC99',
    'red': 'FFC7CE',
}

COLOR_LABEL = {
    'green': 'Yashil',
    'yellow': 'Sariq',
    'orange': "To'q sariq",
    'red': 'Qizil',
}


# ==================== CUSTOM FILTERS ====================

class PsychologicalColorFilter(admin.SimpleListFilter):
    """Psixologik natijalar uchun rang bo'yicha filter"""
    title = 'Kategoriya rangi'
    parameter_name = 'category_color'

    def lookups(self, request, model_admin):
        return [
            ('red', '🔴 Qizil'),
            ('yellow', '🟡 Sariq'),
            ('orange', '🟠 To\'q sariq'),
            ('green', '🟢 Yashil'),
        ]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(
                scale_results__category__color=self.value()
            ).distinct()
        return queryset


class FacultyFilter(admin.SimpleListFilter):
    """Fakultet bo'yicha filter (Result uchun)"""
    title = 'Fakultet'
    parameter_name = 'faculty'

    def lookups(self, request, model_admin):
        from student.models import Student
        faculties = Student.objects.values_list('faculty', flat=True).distinct()
        return [(f, f) for f in faculties if f]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(attempt__student__faculty=self.value())
        return queryset


class GroupFilter(admin.SimpleListFilter):
    """Guruh bo'yicha filter (Result uchun)"""
    title = 'Guruh'
    parameter_name = 'group'

    def lookups(self, request, model_admin):
        from student.models import StudentGroup
        groups = StudentGroup.objects.all()
        return [(g.id, g.group_name) for g in groups]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(attempt__student__group__id=self.value())
        return queryset


# ==================== INLINE ADMINS ====================

class PsychologicalScaleInline(admin.TabularInline):
    """Quiz ichida shkalalarni ko'rsatish"""
    model = PsychologicalScale
    extra = 1
    fields = ('name', 'description', 'order')


class PsychologicalCategoryInline(admin.TabularInline):
    """Shkala ichida kategoriyalarni ko'rsatish"""
    model = PsychologicalCategory
    extra = 1
    fields = ('name', 'min_score', 'max_score', 'color', 'order', 'description')


class OptionInline(admin.TabularInline):
    """Question ichida javoblarni ko'rsatish"""
    model = Option
    extra = 4
    fields = ('option_text', 'is_correct', 'psychological_score', 'order')
    
    def get_formset(self, request, obj=None, **kwargs):
        """Dinamik fieldlar - test turiga qarab"""
        formset = super().get_formset(request, obj, **kwargs)
        
        # Agar quiz mavjud bo'lsa
        if obj and obj.quiz:
            if obj.quiz.is_psychological():
                self.fields = ('option_text', 'psychological_score', 'order')
            else:
                self.fields = ('option_text', 'is_correct', 'order')
        
        return formset


class QuestionInline(admin.TabularInline):
    """Quiz ichida savollarni ko'rsatish (small preview)"""
    model = Question
    extra = 0
    fields = ('question_text', 'score', 'psychological_scale', 'order')
    show_change_link = True


class PsychologicalScaleResultInline(admin.TabularInline):
    """Psixologik natija ichida shkala natijalarini ko'rsatish"""
    model = PsychologicalScaleResult
    extra = 0
    readonly_fields = ('scale', 'total_score', 'category', 'category_color')
    can_delete = False
    
    def category_color(self, obj):
        """Kategoriya rangini ko'rsatish"""
        if obj.category:
            colors = {
                'green': '#10B981',
                'yellow': '#F59E0B',
                'orange': '#F97316',
                'red': '#EF4444',
            }
            color = colors.get(obj.category.color, '#6B7280')
            return format_html(
                '<span style="display:inline-block;width:20px;height:20px;'
                'background:{};border-radius:50%;"></span> {}',
                color,
                obj.category.name
            )
        return '-'
    
    category_color.short_description = 'Kategoriya'


# ==================== MAIN ADMINS ====================

@admin.register(Quiz)
class QuizAdmin(admin.ModelAdmin):
    """Quiz admin"""
    
    list_display = (
        'quiz_icon',
        'title',
        'quiz_type',
        'time_limit',
        'passing_score',
        'questions_count',
        'attempts_count',
        'attempt_limit',
        'is_active',
        'created_at'
    )
    
    list_filter = (
        'quiz_type',
        'is_active',
        'created_at',
    )
    
    search_fields = ('title', 'description')
    
    fieldsets = (
        ('Asosiy Ma\'lumotlar', {
            'fields': ('title', 'description', 'quiz_type', 'is_active', 'attempt_limit')
        }),
        ('Test Sozlamalari', {
            'fields': ('time_limit', 'passing_score'),
            'description': 'passing_score faqat standart testlar uchun ishlatiladi'
        }),
        ('Texnik Ma\'lumotlar', {
            'fields': ('created_by',),
            'classes': ('collapse',)
        }),
    )
    
    inlines = [PsychologicalScaleInline, QuestionInline]
    
    def save_model(self, request, obj, form, change):
        """Yaratuvchini avtomatik o'rnatish"""
        if not change:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def quiz_icon(self, obj):
        """Test turi ikonasi"""
        if obj.is_psychological():
            return format_html('<span style="font-size:20px;">🧠</span>')
        return format_html('<span style="font-size:20px;">📝</span>')
    
    quiz_icon.short_description = ''
    
    def questions_count(self, obj):
        """Savollar soni"""
        count = obj.get_total_questions()
        return format_html(
            '<strong style="color:#4F46E5;">{}</strong> ta',
            count
        )
    
    questions_count.short_description = 'Savollar'
    
    def attempts_count(self, obj):
        """Urinishlar soni"""
        count = obj.attempts.count()
        return format_html(
            '<span style="color:#10B981;">{}</span> urinish',
            count
        )
    
    attempts_count.short_description = 'Urinishlar'


@admin.register(PsychologicalScale)
class PsychologicalScaleAdmin(admin.ModelAdmin):
    """Psixologik shkala admin"""
    
    list_display = ('name', 'quiz', 'categories_count', 'order')
    list_filter = ('quiz',)
    search_fields = ('name', 'description')
    
    inlines = [PsychologicalCategoryInline]
    
    def categories_count(self, obj):
        """Kategoriyalar soni"""
        return obj.categories.count()
    
    categories_count.short_description = 'Kategoriyalar'


@admin.register(PsychologicalCategory)
class PsychologicalCategoryAdmin(admin.ModelAdmin):
    """Kategoriya admin"""
    
    list_display = ('name', 'scale', 'score_range', 'color_display', 'order')
    list_filter = ('scale', 'color')
    search_fields = ('name', 'description')
    
    def score_range(self, obj):
        """Ball oralig'i"""
        return f"{obj.min_score} - {obj.max_score}"
    
    score_range.short_description = 'Ball oralig\'i'
    
    def color_display(self, obj):
        """Rang ko'rsatish"""
        colors = {
            'green': '#10B981',
            'yellow': '#F59E0B',
            'orange': '#F97316',
            'red': '#EF4444',
        }
        color = colors.get(obj.color, '#6B7280')
        return format_html(
            '<span style="display:inline-block;width:50px;height:20px;'
            'background:{};border-radius:4px;"></span>',
            color
        )
    
    color_display.short_description = 'Rang'


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    """Savol admin"""
    
    list_display = (
        'quiz',
        'question_preview',
        'quiz_type_display',
        'score',
        'psychological_scale',
        'options_count',
        'order'
    )
    
    list_filter = (
        'quiz__quiz_type',
        'quiz',
        'psychological_scale',
    )
    
    search_fields = ('question_text',)
    
    inlines = [OptionInline]
    
    fieldsets = (
        ('Savol', {
            'fields': ('quiz', 'question_text', 'order')
        }),
        ('Ball va Shkala', {
            'fields': ('score', 'psychological_scale'),
            'description': (
                'Standart testda faqat "score" ishlatiladi. '
                'Psixologik testda "psychological_scale" tanlash majburiy.'
            )
        }),
    )
    
    def question_preview(self, obj):
        """Savol previewi"""
        text = obj.question_text[:100]
        if len(obj.question_text) > 100:
            text += '...'
        return text
    
    question_preview.short_description = 'Savol'
    
    def quiz_type_display(self, obj):
        """Test turi"""
        if obj.quiz.is_psychological():
            return format_html('<span style="color:#7C3AED;">🧠 Psixologik</span>')
        return format_html('<span style="color:#4F46E5;">📝 Standart</span>')
    
    quiz_type_display.short_description = 'Test turi'
    
    def options_count(self, obj):
        """Javoblar soni"""
        return obj.options.count()
    
    options_count.short_description = 'Javoblar'


@admin.register(Option)
class OptionAdmin(admin.ModelAdmin):
    """Javob varianti admin"""
    
    list_display = (
        'option_text',
        'question',
        'quiz_type_display',
        'is_correct',
        'psychological_score',
        'order'
    )
    
    list_filter = (
        'question__quiz__quiz_type',
        'is_correct',
    )
    
    search_fields = ('option_text', 'question__question_text')
    
    def quiz_type_display(self, obj):
        """Test turi"""
        if obj.question.quiz.is_psychological():
            return format_html(
                '<span style="color:#7C3AED;">🧠 Psixologik ({} ball)</span>',
                obj.psychological_score
            )
        
        status = 'To\'g\'ri' if obj.is_correct else 'Noto\'g\'ri'
        color = '#10B981' if obj.is_correct else '#EF4444'
        return format_html(
            '<span style="color:{};">📝 {} </span>',
            color,
            status
        )
    
    quiz_type_display.short_description = 'Holat'


@admin.register(QuestionText)
class QuestionTextAdmin(admin.ModelAdmin):
    """Bulk upload admin"""
    
    list_display = ('quiz', 'status_display', 'score', 'created_at')
    list_filter = ('is_processed', 'quiz')
    readonly_fields = ('is_processed',)
    
    def status_display(self, obj):
        """Holat"""
        if obj.is_processed:
            return format_html(
                '<span style="color:#10B981;font-weight:bold;">✓ Qayta ishlangan</span>'
            )
        return format_html(
            '<span style="color:#F59E0B;font-weight:bold;">⏳ Kutilmoqda</span>'
        )
    
    status_display.short_description = 'Holat'


@admin.register(QuizAttempt)
class QuizAttemptAdmin(admin.ModelAdmin):
    """Urinish admin"""
    
    list_display = (
        'student',
        'quiz',
        'quiz_type_display',
        'status',
        'started_at',
        'completed_at',
        'time_display',
        'view_result'
    )
    
    list_filter = (
        'status',
        'quiz__quiz_type',
        'quiz',
        'started_at',
    )
    
    search_fields = (
        'student__student_name',
        'student__student_id_number',
        'quiz__title'
    )
    
    readonly_fields = (
        'started_at',
        'completed_at',
        'time_taken'
    )
    
    def quiz_type_display(self, obj):
        """Test turi"""
        if obj.quiz.is_psychological():
            return '🧠 Psixologik'
        return '📝 Standart'
    
    quiz_type_display.short_description = 'Tur'
    
    def time_display(self, obj):
        """Vaqt ko'rsatish"""
        if obj.time_taken:
            minutes = obj.time_taken // 60
            seconds = obj.time_taken % 60
            return f"{minutes}m {seconds}s"
        return '-'
    
    time_display.short_description = 'Vaqt'
    
    def view_result(self, obj):
        """Natijani ko'rish"""
        if obj.status == 'completed':
            if obj.quiz.is_standard() and hasattr(obj, 'result'):
                return format_html(
                    '<a href="/admin/main/result/{}/change/" '
                    'style="color:#4F46E5;font-weight:bold;">📊 Natija</a>',
                    obj.result.id
                )
            elif obj.quiz.is_psychological() and hasattr(obj, 'psychological_result'):
                return format_html(
                    '<a href="/admin/main/psychologicalresult/{}/change/" '
                    'style="color:#7C3AED;font-weight:bold;">🧠 Natija</a>',
                    obj.psychological_result.id
                )
        return '-'
    
    view_result.short_description = 'Natija'


# ==================== RESULT ADMIN ====================

@admin.register(Result)
class ResultAdmin(admin.ModelAdmin):
    """Standart test natijasi admin"""
    
    list_display = (
        'student_name',
        'student_faculty',
        'student_group',
        'quiz',
        'percentage_display',
        'grade_display',
        'passed_display',
        'created_at'
    )
    
    list_filter = (
        'passed',
        'attempt__quiz',
        FacultyFilter,
        GroupFilter,
        'created_at',
    )
    
    search_fields = (
        'attempt__student__student_name',
        'attempt__quiz__title'
    )
    
    readonly_fields = (
        'attempt',
        'total_questions',
        'correct_answers',
        'wrong_answers',
        'unanswered',
        'total_score',
        'max_score',
        'percentage',
        'passed'
    )

    actions = ['export_results_excel']
    
    def student_name(self, obj):
        """Talaba ismi"""
        return obj.attempt.student.student_name
    
    student_name.short_description = 'Talaba'

    def student_faculty(self, obj):
        """Talaba fakulteti"""
        return obj.attempt.student.faculty or '-'

    student_faculty.short_description = 'Fakultet'

    def student_group(self, obj):
        """Talaba guruhi"""
        group = obj.attempt.student.group
        return group.group_name if group else '-'

    student_group.short_description = 'Guruh'
    
    def quiz(self, obj):
        """Test nomi"""
        return obj.attempt.quiz.title
    
    quiz.short_description = 'Test'
    
    def percentage_display(self, obj):
        """Foiz ko'rsatish"""
        color = '#10B981' if obj.passed else '#EF4444'
        return format_html(
            '<strong style="color:{}; font-size:16px;">{:.1f}%</strong>',
            color,
            obj.percentage
        )
    
    percentage_display.short_description = 'Natija'
    
    def grade_display(self, obj):
        """Baho"""
        grade = obj.get_grade()
        colors = {5: '#10B981', 4: '#3B82F6', 3: '#F59E0B', 2: '#EF4444'}
        return format_html(
            '<span style="background:{}; color:white; padding:4px 12px; '
            'border-radius:12px; font-weight:bold;">{}</span>',
            colors.get(grade, '#6B7280'),
            grade
        )
    
    grade_display.short_description = 'Baho'
    
    def passed_display(self, obj):
        """O'tdi/O'tmadi"""
        if obj.passed:
            return format_html(
                '<span style="color:#10B981;font-weight:bold;">✓ O\'tdi</span>'
            )
        return format_html(
            '<span style="color:#EF4444;font-weight:bold;">✗ O\'tmadi</span>'
        )
    
    passed_display.short_description = 'Holat'

    @admin.action(description='📥 Tanlangan natijalarni Excel yuklab olish')
    def export_results_excel(self, request, queryset):
        """Tanlangan standart test natijalarini Excel formatda yuklab olish"""
        if not OPENPYXL_AVAILABLE:
            self.message_user(request, "openpyxl o'rnatilmagan! pip install openpyxl", level='error')
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Natijalar"

        headers = [
            '№', 'Talaba ismi', 'Fakultet', 'Guruh', 'Test nomi',
            'Jami savollar', "To'g'ri", "Noto'g'ri", 'Javobsiz',
            'Ball', 'Max ball', 'Foiz (%)', 'Baho', 'Holat', 'Sana'
        ]

        ws.append(headers)
        _style_header_row(ws, 1, len(headers), fill_color="4F46E5")

        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for idx, result in enumerate(queryset.select_related(
            'attempt__student__group', 'attempt__quiz'
        ), start=1):
            student = result.attempt.student
            group = student.group
            row = [
                idx,
                student.student_name or '-',
                student.faculty or '-',
                group.group_name if group else '-',
                result.attempt.quiz.title,
                result.total_questions,
                result.correct_answers,
                result.wrong_answers,
                result.unanswered,
                result.total_score,
                result.max_score,
                float(result.percentage),
                result.get_grade(),
                "O'tdi" if result.passed else "O'tmadi",
                result.created_at.strftime('%Y-%m-%d %H:%M') if result.created_at else '-',
            ]
            ws.append(row)

            # Rangni belgilash
            row_num = ws.max_row
            fill = pass_fill if result.passed else fail_fill
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill

        _auto_column_width(ws)
        ws.freeze_panes = "A2"

        return _make_excel_response(wb, "natijalar.xlsx")


# ==================== PSYCHOLOGICAL RESULT ADMIN ====================

@admin.register(PsychologicalResult)
class PsychologicalResultAdmin(admin.ModelAdmin):
    """Psixologik test natijasi admin"""

    list_display = (
        'student_name',
        'student_faculty',
        'student_group',
        'quiz',
        'answered_display',
        'scales_summary',
        'created_at'
    )

    list_filter = (
        'attempt__quiz',
        FacultyFilter,
        GroupFilter,
        PsychologicalColorFilter,
        'created_at',
    )

    search_fields = (
        'attempt__student__student_name',
        'attempt__quiz__title'
    )

    readonly_fields = (
        'attempt',
        'total_questions',
        'answered_questions',
        'unanswered'
    )

    inlines = [PsychologicalScaleResultInline]

    actions = ['export_psychological_excel', 'export_single_psychological_excel']

    def student_name(self, obj):
        """Talaba ismi"""
        return obj.attempt.student.student_name

    student_name.short_description = 'Talaba'

    def student_faculty(self, obj):
        """Talaba fakulteti"""
        return obj.attempt.student.faculty or '-'

    student_faculty.short_description = 'Fakultet'

    def student_group(self, obj):
        """Talaba guruhi"""
        group = obj.attempt.student.group
        return group.group_name if group else '-'

    student_group.short_description = 'Guruh'

    def quiz(self, obj):
        """Test nomi"""
        return obj.attempt.quiz.title

    quiz.short_description = 'Test'

    def answered_display(self, obj):
        """Javob berilgan savollar"""
        return format_html(
            '<strong>{}</strong> / {} ta',
            obj.answered_questions,
            obj.total_questions
        )

    answered_display.short_description = 'Javoblar'

    def scales_summary(self, obj):
        """Shkalalar qisqacha (ranglar bilan)"""
        results = obj.scale_results.select_related('scale', 'category')
        if not results:
            return '-'

        color_html = {
            'green': '#10B981',
            'yellow': '#F59E0B',
            'orange': '#F97316',
            'red': '#EF4444',
        }

        parts = []
        for sr in results:
            if sr.category:
                c = color_html.get(sr.category.color, '#6B7280')
                parts.append(format_html(
                    '<span style="color:{};">● {}: {} ({})</span>',
                    c, sr.scale.name, sr.total_score, sr.category.name
                ))

        if not parts:
            return '-'

        # format_html_join not used to keep it simple
        from django.utils.safestring import mark_safe
        return mark_safe(' &nbsp;|&nbsp; '.join(parts))

    scales_summary.short_description = 'Natijalar'

    @admin.action(description='📥 Tanlangan psixologik natijalarni Excel yuklab olish')
    def export_psychological_excel(self, request, queryset):
        """Ko'p talabaning psixologik natijalarini Excel formatda yuklab olish"""
        if not OPENPYXL_AVAILABLE:
            self.message_user(request, "openpyxl o'rnatilmagan! pip install openpyxl", level='error')
            return

        # Barcha shkalalarni olish
        all_scales = []
        for result in queryset.prefetch_related('scale_results__scale'):
            for sr in result.scale_results.all():
                scale_name = sr.scale.name
                if scale_name not in all_scales:
                    all_scales.append(scale_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Psixologik natijalar"

        # Headerlar
        base_headers = ['№', 'Talaba ismi', 'Fakultet', 'Guruh', 'Test nomi',
                        'Jami savollar', 'Javob berilgan', 'Javobsiz', 'Sana']
        scale_headers = []
        for s in all_scales:
            scale_headers.append(f"{s} (ball)")
            scale_headers.append(f"{s} (kategoriya)")
            scale_headers.append(f"{s} (rang)")

        headers = base_headers + scale_headers
        ws.append(headers)
        _style_header_row(ws, 1, len(headers), fill_color="7C3AED")

        for idx, result in enumerate(queryset.select_related(
            'attempt__student__group', 'attempt__quiz'
        ).prefetch_related('scale_results__scale', 'scale_results__category'), start=1):
            student = result.attempt.student
            group = student.group

            row = [
                idx,
                student.student_name or '-',
                student.faculty or '-',
                group.group_name if group else '-',
                result.attempt.quiz.title,
                result.total_questions,
                result.answered_questions,
                result.unanswered,
                result.created_at.strftime('%Y-%m-%d %H:%M') if result.created_at else '-',
            ]

            # Shkala natijalarini map
            scale_map = {}
            for sr in result.scale_results.all():
                scale_map[sr.scale.name] = sr

            for scale_name in all_scales:
                sr = scale_map.get(scale_name)
                if sr:
                    row.append(sr.total_score)
                    row.append(sr.category.name if sr.category else '-')
                    row.append(COLOR_LABEL.get(sr.category.color, '-') if sr.category else '-')
                else:
                    row.extend(['-', '-', '-'])

            ws.append(row)

            # Rang ranglash - birinchi rang categoriyasiga qarab
            row_num = ws.max_row
            first_color = None
            for sr in result.scale_results.all():
                if sr.category:
                    first_color = sr.category.color
                    break
            if first_color and first_color in COLOR_MAP:
                fill = PatternFill(
                    start_color=COLOR_MAP[first_color],
                    end_color=COLOR_MAP[first_color],
                    fill_type="solid"
                )
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = fill

        _auto_column_width(ws)
        ws.freeze_panes = "A2"

        return _make_excel_response(wb, "psixologik_natijalar.xlsx")

    @admin.action(description='📋 Bitta talabaning batafsil psixologik natijasini yuklab olish')
    def export_single_psychological_excel(self, request, queryset):
        """Bitta talabaning batafsil psixologik natijasini yuklab olish"""
        if not OPENPYXL_AVAILABLE:
            self.message_user(request, "openpyxl o'rnatilmagan! pip install openpyxl", level='error')
            return

        if queryset.count() != 1:
            self.message_user(
                request,
                "Bu action faqat BITTA talaba uchun ishlaydi. Iltimos, faqat bitta natija tanlang.",
                level='warning'
            )
            return

        result = queryset.select_related(
            'attempt__student__group', 'attempt__quiz'
        ).prefetch_related(
            'scale_results__scale',
            'scale_results__category'
        ).first()

        student = result.attempt.student
        group = student.group

        wb = openpyxl.Workbook()

        # ---- 1-varaq: Umumiy ma'lumot ----
        ws1 = wb.active
        ws1.title = "Umumiy"

        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        info_rows = [
            ("Talaba ismi", student.student_name or '-'),
            ("Talaba ID", student.student_id_number or '-'),
            ("HEMIS ID", student.hemis_id or '-'),
            ("Fakultet", student.faculty or '-'),
            ("Guruh", group.group_name if group else '-'),
            ("Kurs", student.level or '-'),
            ("Semestr", student.semester or '-'),
            ("Test nomi", result.attempt.quiz.title),
            ("Test sanasi", result.created_at.strftime('%Y-%m-%d %H:%M') if result.created_at else '-'),
            ("Jami savollar", result.total_questions),
            ("Javob berilgan", result.answered_questions),
            ("Javobsiz", result.unanswered),
        ]

        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 45

        for r_idx, (label, value) in enumerate(info_rows, start=1):
            ws1.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
            ws1.cell(row=r_idx, column=2, value=value)
            if r_idx % 2 == 0:
                light_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                ws1.cell(row=r_idx, column=1).fill = light_fill
                ws1.cell(row=r_idx, column=2).fill = light_fill

        # ---- 2-varaq: Shkala natijalari ----
        ws2 = wb.create_sheet(title="Shkala natijalari")

        scale_headers = ['Shkala nomi', 'Ball', 'Kategoriya', 'Rang', "Kategoriya tavsifi"]
        ws2.append(scale_headers)
        _style_header_row(ws2, 1, len(scale_headers), fill_color="7C3AED")

        for sr in result.scale_results.select_related('scale', 'category'):
            color_label = COLOR_LABEL.get(sr.category.color, '-') if sr.category else '-'
            desc = sr.category.description if sr.category else '-'
            row_data = [
                sr.scale.name,
                sr.total_score,
                sr.category.name if sr.category else '-',
                color_label,
                desc,
            ]
            ws2.append(row_data)

            # Rang bo'yicha fill
            row_num = ws2.max_row
            if sr.category and sr.category.color in COLOR_MAP:
                fill = PatternFill(
                    start_color=COLOR_MAP[sr.category.color],
                    end_color=COLOR_MAP[sr.category.color],
                    fill_type="solid"
                )
                for col in range(1, len(scale_headers) + 1):
                    ws2.cell(row=row_num, column=col).fill = fill

        _auto_column_width(ws2)
        ws2.freeze_panes = "A2"

        # Fayl nomi
        safe_name = (student.student_name or 'talaba').replace(' ', '_')
        filename = f"{safe_name}_psixologik_natija.xlsx"

        return _make_excel_response(wb, filename)


# ==================== USER RESPONSE ====================

@admin.register(UserResponse)
class UserResponseAdmin(admin.ModelAdmin):
    """Javob admin (faqat ko'rish uchun)"""
    
    list_display = ('student_name', 'quiz_name', 'question_preview', 'response_display', 'answered_at')
    
    list_filter = ('attempt__quiz__quiz_type', 'is_correct', 'answered_at')
    
    search_fields = ('attempt__student__student_name', 'question__question_text')
    
    readonly_fields = ('attempt', 'question', 'selected_option', 'is_correct', 'earned_score', 'answered_at')
    
    def student_name(self, obj):
        return obj.attempt.student.student_name
    
    student_name.short_description = 'Talaba'
    
    def quiz_name(self, obj):
        return obj.attempt.quiz.title
    
    quiz_name.short_description = 'Test'
    
    def question_preview(self, obj):
        text = obj.question.question_text[:50]
        if len(obj.question.question_text) > 50:
            text += '...'
        return text
    
    question_preview.short_description = 'Savol'
    
    def response_display(self, obj):
        """Javob ko'rsatish"""
        if not obj.selected_option:
            return format_html('<span style="color:#6B7280;">-</span>')
        
        if obj.attempt.quiz.is_psychological():
            return format_html(
                '<span style="color:#7C3AED;">{} ({} ball)</span>',
                obj.selected_option.option_text[:30],
                obj.earned_score
            )
        else:
            color = '#10B981' if obj.is_correct else '#EF4444'
            icon = '✓' if obj.is_correct else '✗'
            return format_html(
                '<span style="color:{};font-weight:bold;">{} {}</span>',
                color,
                icon,
                obj.selected_option.option_text[:30]
            )
    
    response_display.short_description = 'Javob'